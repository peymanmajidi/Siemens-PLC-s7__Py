# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
from PIL import Image


  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('excel.xlsx') 
  
# create the sheet object 
sheet = wb.active 
  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 20
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Email id"
    sheet.cell(row=1, column=3).value = "Address"
  
  
# Function to set focus (cursor) 
  
  
# Function to set focus 
def focus5(event): 
    # set focus on the email_id_field box 
    email_id_field.focus_set() 
  
  
# Function to set focus 
def focus6(event): 
    # set focus on the address_field box 
    address_field.focus_set() 
  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    email_id_field.delete(0, END) 
    address_field.delete(0, END) 
  
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = email_id_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = address_field.get() 
  
        # save the file 
        wb.save('excel.xlsx') 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("registration form") 

  
    # set the configuration of GUI window 
    root.geometry("500x300") 

    root.grid_rowconfigure(5, weight=1)
    root.grid_columnconfigure(4, weight=1)
  
    excel() 

  
    # create a Form label 
    heading = Label(root, text="Please enter your contact info", bg="light green") 
  
    # create a Name label 
    name = Label(root, text="Name", bg="light green") 
    
    email_id = Label(root, text="Email id", bg="light green") 
  
    # create a address label 
    address = Label(root, text="Address", bg="light green") 

  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    email_id.grid(row=2, column=0) 
    address.grid(row=3, column=0) 
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    email_id_field = Entry(root) 
    address_field = Entry(root) 
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 

    # then call the focus6 function 
    email_id_field.bind("<Return>", focus6) 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=1, ipadx="100") 
    email_id_field.grid(row=2, column=1, ipadx="100") 
    address_field.grid(row=3, column=1, ipadx="100") 
  
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                             command=insert) 
    submit.grid(row=4, column=1,columnspan=2, sticky="nsew") 
  
    # start the GUI 
    root.mainloop()